"""Headless Excel bridge (needs ``openpyxl``; no Excel install, no xlwings).

* :func:`write_input_template` - a clean, labelled input workbook whose cells
  map 1:1 onto the project schema.
* :func:`read_project` - parse that workbook (or one saved from it) back into a
  project ``dict`` for :class:`pumpsizer.Project`.
* :func:`write_results` - a multi-sheet results workbook (summary, curve data,
  EPANET block, selection, surge).
* :func:`read_legacy_workbook` - best-effort mapping of the original
  ``Pump Sizing.xlsx`` Input sheet (cell refs pinned in code).

Later a one-line xlwings macro / button can call ``Project.from_dict(
read_project(path)).run()`` and ``write_results(...)`` - this module keeps the
engine usable with nothing but Python.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise ImportError("the Excel bridge needs openpyxl: pip install openpyxl "
                      "(or pip install pumpsizer[excel])") from exc

from .project import Project, ProjectResults
from .report import text_report

_HDR = Font(bold=True, size=12)
_KEY = Font(bold=True)
_FILL = PatternFill("solid", fgColor="DDE7F0")
_INPUT_FILL = PatternFill("solid", fgColor="FFF6D5")

# ---------------------------------------------------------------------------
# input template  (Sheet "Input": section | key | value | unit/notes)
# ---------------------------------------------------------------------------
# (section, dotted-schema-key, default, unit/hint)
_TEMPLATE_ROWS: list[tuple[str, str, Any, str]] = [
    ("project", "project.name", "My Pumping Station", ""),
    ("project", "project.description", "", ""),
    ("fluid", "fluid.temperature_c", 30, "degC"),
    ("fluid", "fluid.altitude_m", 0, "m above MSL"),
    ("pipe", "pipe.material", "ductile_iron", "ductile_iron|steel|grp|hdpe|upvc"),
    ("pipe", "pipe.series", "", "SDR name for hdpe/upvc"),
    ("pipe", "pipe.headloss_method", "DW", "DW|HW"),
    ("pipe", "pipe.pressure_class_pn", 16, "bar (for surge rating)"),
    ("flow", "flow.total_demand_lps", 300, "l/s"),
    ("flow", "flow.duty_pumps", 1, "count"),
    ("flow", "flow.standby_pumps", 1, "count"),
    ("levels", "levels.reservoir_hwl_m", 32, "m"),
    ("levels", "levels.reservoir_bwl_m", 25, "m"),
    ("levels", "levels.sump_hwl_m", 15, "m"),
    ("levels", "levels.sump_bwl_m", 8, "m"),
    ("levels", "levels.pump_centreline_m", 8, "m (flooded if <= sump BWL)"),
    ("suction", "suction.npsh_safety_margin_m", 0.5, "m"),
    ("pump", "pump.source", "synthetic", "synthetic|single_point|points|catalogue"),
    ("pump", "pump.name", "WS-300", ""),
    ("pump", "pump.shutoff_ratio", 1.20, "shutoff/design head"),
    ("pump", "pump.bep_efficiency_pct", 84, "%"),
    ("pump", "pump.catalogue_path", "", "file/dir when source=catalogue"),
    ("pump", "pump.model", "", "force a catalogue model (optional)"),
    ("control", "control.arrangement", "single", "single|parallel|series"),
    ("control", "control.vfd", False, "TRUE/FALSE"),
    ("control", "control.vfd_min_speed_pct", 70, "%"),
    ("control", "control.vfd_target_flow_lps", "", "l/s (blank = total demand)"),
    ("motor", "motor.poles", 2, "2|4|6|8"),
    ("motor", "motor.ie_class", "IE3", "IE1|IE2|IE3|IE4"),
    ("motor", "motor.rating_margin_pct", 15, "%"),
    ("motor", "motor.sizing_basis", "operating_point", "operating_point|non_overloading"),
    ("energy", "energy.hours_per_day", 20, "h"),
    ("energy", "energy.tariff_per_kwh", 0.12, "currency/kWh"),
    ("energy", "energy.life_cycle_years", 20, "years"),
    ("energy", "energy.discount_rate", 0.08, "fraction"),
    ("energy", "energy.capital_cost", 0, "currency"),
    ("water_hammer", "water_hammer.enabled", True, "TRUE/FALSE"),
    ("water_hammer", "water_hammer.closure_time_s", "", "s (blank = rapid)"),
    ("water_hammer", "water_hammer.allowable_max_head_m", 150, "m"),
    ("epanet", "epanet.flow_units", "LPS", "LPS|LPM|MLD|CMH|CMD|GPM"),
    ("epanet", "epanet.pump_id", "PMP1", ""),
    ("epanet", "epanet.from_node", "SUMP", ""),
    ("epanet", "epanet.to_node", "PUMP_OUT", ""),
    ("epanet", "epanet.head_points", 3, "3 = A-B*Q^C refit; >3 = multipoint"),
]

_SEGMENT_COLS = ["name", "group", "length_m", "dn", "diameter_mm"]
_DEFAULT_SEGMENTS = [
    ["suction_header", "suction", 25, 700, ""],
    ["pump_suction", "suction", 10, 700, ""],
    ["pump_discharge", "discharge", 15, 400, ""],
    ["rising_main", "discharge", 500, 400, ""],
]
_DEFAULT_FITTINGS = [
    ["suction", "entrance_bellmouth", 1], ["suction", "reducer", 1],
    ["suction", "gate_valve", 4], ["suction", "bend_90", 4], ["suction", "tee", 1],
    ["discharge", "tee", 1], ["discharge", "enlarger", 1], ["discharge", "bend_45", 4],
    ["discharge", "bend_90", 4], ["discharge", "bend_22_5", 6], ["discharge", "gate_valve", 1],
    ["discharge", "butterfly_valve", 1], ["discharge", "non_return_valve", 1],
    ["discharge", "exit_sharp", 1],
]
_FIT_REF = {"suction": "pump_suction", "discharge": "pump_discharge"}


def _autosize(ws) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 3, 60)


def write_input_template(path: str | Path) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input"
    ws.append(["section", "key", "value", "unit / hint"])
    for c in ws[1]:
        c.font = _KEY
        c.fill = _FILL
    last_section = None
    for section, key, default, hint in _TEMPLATE_ROWS:
        r = ws.max_row + 1
        ws.cell(r, 1, section if section != last_section else "")
        ws.cell(r, 2, key)                      # full dotted key, unambiguous
        vcell = ws.cell(r, 3, default)
        vcell.fill = _INPUT_FILL
        ws.cell(r, 4, hint)
        last_section = section
    ws.freeze_panes = "A2"
    _autosize(ws)

    seg = wb.create_sheet("Segments")
    seg.append(_SEGMENT_COLS)
    for c in seg[1]:
        c.font = _KEY
        c.fill = _FILL
    for row in _DEFAULT_SEGMENTS:
        seg.append(row)
    _autosize(seg)

    fit = wb.create_sheet("Fittings")
    fit.append(["group", "fitting", "quantity"])
    for c in fit[1]:
        c.font = _KEY
        c.fill = _FILL
    for row in _DEFAULT_FITTINGS:
        fit.append(row)
    _autosize(fit)

    pts = wb.create_sheet("PumpCurve")
    pts.append(["flow_lps", "head_m", "efficiency_pct", "npshr_m"])
    for c in pts[1]:
        c.font = _KEY
        c.fill = _FILL
    pts.append(["# fill these rows only when Input!pump.source = points", "", "", ""])
    _autosize(pts)

    wb.save(path)
    return str(path)


# ---------------------------------------------------------------------------
# read
# ---------------------------------------------------------------------------
def _coerce(v: Any) -> Any:
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None
        if s.lower() in ("true", "yes"):
            return True
        if s.lower() in ("false", "no"):
            return False
        return s
    return v


def _nest(flat: dict[str, Any]) -> dict:
    out: dict = {}
    for dotted, val in flat.items():
        if val is None:
            continue
        cur = out
        parts = dotted.split(".")
        for p in parts[:-1]:
            cur = cur.setdefault(p, {})
        cur[parts[-1]] = val
    return out


def read_project(path: str | Path) -> dict:
    """Parse a template-shaped workbook into a project dict."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Input"]
    flat: dict[str, Any] = {}
    known = {k for _, k, _, _ in _TEMPLATE_ROWS}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[1] in (None, ""):
            continue
        key = str(row[1]).strip()
        if "." not in key:                      # tolerate a legacy short-key sheet
            key = next((k for k in known if k.split(".", 1)[1] == key), key)
        if key in known:
            flat[key] = _coerce(row[2])
    data = _nest(flat)

    if "Segments" in wb.sheetnames:
        segs = []
        sh = wb["Segments"]
        for row in sh.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            d = dict(zip(_SEGMENT_COLS, row))
            seg: dict[str, Any] = {"name": str(d["name"]), "group": str(d.get("group") or "discharge"),
                                   "length_m": float(d["length_m"])}
            if d.get("diameter_mm"):
                seg["diameter_mm"] = float(d["diameter_mm"])
            elif d.get("dn"):
                seg["dn"] = float(d["dn"])
            segs.append(seg)
        if segs:
            data["segments"] = segs

    if "Fittings" in wb.sheetnames:
        fittings: dict[str, dict] = {}
        for row in wb["Fittings"].iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not row[1]:
                continue
            fittings.setdefault(str(row[0]), {})[str(row[1])] = float(row[2] or 0)
        if fittings:
            data["fittings"] = fittings
            data.setdefault("fitting_reference", _FIT_REF)

    if "PumpCurve" in wb.sheetnames and data.get("pump", {}).get("source") == "points":
        q, h, e, n = [], [], [], []
        for row in wb["PumpCurve"].iter_rows(min_row=2, values_only=True):
            if not row or not isinstance(row[0], (int, float)):
                continue
            q.append(float(row[0]))
            h.append(float(row[1]))
            if row[2] is not None:
                e.append(float(row[2]))
            if len(row) > 3 and row[3] is not None:
                n.append(float(row[3]))
        if q:
            data["pump"]["curve_points"] = {"flow_lps": q, "head_m": h}
            if len(e) == len(q):
                data["pump"]["efficiency_points"] = {"flow_lps": q, "value_pct": e}
            if len(n) == len(q):
                data["pump"]["npshr_points"] = {"flow_lps": q, "value_m": n}
    return data


# ---- legacy Pump Sizing.xlsx --------------------------------------------
def read_legacy_workbook(path: str | Path) -> dict:
    """Best-effort mapping of the original ``Pump Sizing.xlsx`` Input sheet.
    Cell references are pinned to that layout - re-check if the sheet changes."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Input"]

    def D(row: int):
        return ws.cell(row=row, column=4).value

    material_map = {1: "ductile_iron", 2: "grp", 3: "hdpe", 4: "steel", 5: "upvc"}
    data: dict[str, Any] = {
        "project": {"name": str(ws["A1"].value or "Legacy workbook import").split(":")[-1].strip()},
        "fluid": {"temperature_c": D(24) or 20},
        "pipe": {"headloss_method": "DW"},
        "flow": {"total_demand_lps": D(11), "duty_pumps": int(D(12) or 1)},
        "levels": {
            "reservoir_hwl_m": D(42), "reservoir_bwl_m": D(43),
            "sump_hwl_m": D(44), "sump_bwl_m": D(45),
            "pump_centreline_m": D(46) if D(46) is not None else D(45),
        },
        "control": {"vfd": (D(13) == 2), "vfd_min_speed_pct": (D(14) or 0) * 100},
        "segments": [
            {"name": "suction_header", "group": "suction", "length_m": D(15), "dn": None},
            {"name": "pump_suction", "group": "suction", "length_m": D(16), "dn": None},
            {"name": "pump_discharge", "group": "discharge", "length_m": D(17), "dn": None},
            {"name": "rising_main", "group": "discharge", "length_m": D(18), "dn": None},
        ],
        "fittings": {
            "suction": {"entrance_bellmouth": D(27) or 0, "reducer": D(28) or 0,
                        "gate_valve": D(29) or 0, "bend_90": D(30) or 0, "tee": D(31) or 0},
            "discharge": {"tee": D(33) or 0, "enlarger": D(34) or 0, "bend_45": D(35) or 0,
                          "bend_90": D(36) or 0, "bend_22_5": D(37) or 0, "gate_valve": D(38) or 0,
                          "butterfly_valve": D(39) or 0, "non_return_valve": D(40) or 0,
                          "exit_sharp": D(41) or 0}},
        "fitting_reference": _FIT_REF,
        "autosize_diameter": True,
        "pump": {"source": "synthetic"},
        "water_hammer": {"enabled": True},
    }
    try:
        data["pipe"]["material"] = material_map.get(int(wb["Material"]["C3"].value or 1), "ductile_iron")
    except Exception:
        data["pipe"]["material"] = "ductile_iron"
    # drop empty leaves
    for seg in data["segments"]:
        seg.pop("dn", None)
    return data


# ---------------------------------------------------------------------------
# write results
# ---------------------------------------------------------------------------
def write_results(res: ProjectResults, path: str | Path, *, include_report: bool = True) -> str:
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["PUMP SIZING RESULT", res.project_name])
    ws["A1"].font = _HDR
    op = res.operating_point

    def block(title: str, rows: list[tuple[str, Any]]):
        ws.append([])
        r = ws.max_row + 1
        ws.cell(r, 1, title).font = _KEY
        ws.cell(r, 1).fill = _FILL
        for k, v in rows:
            ws.append([k, v])

    block("Operating point", [
        ("Flow (l/s)", round(op.flow_lps, 2)),
        ("Head (m)", round(op.head_m, 2)),
        ("Running pumps", op.n_pumps),
        ("Flow per pump (l/s)", round(op.flow_per_pump_m3s * 1000, 2)),
        ("Pump efficiency (%)", None if op.efficiency_pct != op.efficiency_pct else round(op.efficiency_pct, 1)),
        ("Hydraulic power total (kW)", round(op.hydraulic_power_kw, 2)),
        ("Shaft power total (kW)", round(op.shaft_power_kw, 2)),
        ("VFD speed (%)", round(op.speed_ratio * 100, 1)),
    ])
    block("System head", [
        ("Design system head (m)", round(res.design_system_head_m, 2)),
        ("Duty per pump", f"{res.duty_flow_per_pump_m3s*1000:.1f} l/s @ {res.duty_head_m:.2f} m"),
    ] + [(f"H_sys [{k}] (m)", round(v.head(res.duty_flow_per_pump_m3s * op.n_pumps), 2))
         for k, v in res.system_set.as_dict().items()])
    n = res.npsh
    block("NPSH", [("NPSH available (m)", n.npsh_available_m),
                   ("NPSH required (m)", n.npsh_required_m),
                   ("Margin (m)", n.margin_m), ("Safe", n.safe)])
    m = res.motor
    block("Motor", [("Shaft power one pump (kW)", round(m.shaft_power_kw, 2)),
                    ("Rated (kW)", m.rated_kw), ("Poles", m.poles), ("IE class", m.ie_class),
                    ("Motor efficiency (%)", round(m.motor_efficiency_pct, 1)),
                    ("Electrical input one pump (kW)", round(m.input_electrical_kw, 2))])
    if res.energy:
        block("Energy", [(k, v) for k, v in res.energy.items() if not isinstance(v, dict)])
    for wn in res.warnings:
        ws.append(["! warning", wn])
    _autosize(ws)

    # curve data
    cs = wb.create_sheet("Curves")
    cs.append(["flow_lps", "system_max_used_m", "system_min_new_m", "pump_head_m",
               "pump_eff_pct", "pump_npshr_m"])
    for c in cs[1]:
        c.font = _KEY
    import numpy as np
    qmax = max(res.pump.max_flow() * max(op.n_pumps, 1), res.duty_flow_per_pump_m3s * 1.4)
    for q in np.linspace(1e-4, qmax, 40):
        cs.append([round(q * 1000, 2),
                   round(float(res.system_set.max_static_used.head(q)), 3),
                   round(float(res.system_set.min_static_new.head(q)), 3),
                   round(float(res.pump.head(q)), 3),
                   None if res.pump.eff_pts is None else round(float(res.pump.efficiency(q)), 2),
                   None if res.pump.npshr_pts is None else round(float(res.pump.npshr(q)), 3)])
    _autosize(cs)

    es = wb.create_sheet("EPANET")
    for i, line in enumerate(res.epanet_export.full_snippet().splitlines(), 1):
        es.cell(i, 1, line)
    es.column_dimensions["A"].width = 70

    if res.selection:
        sl = wb.create_sheet("Selection")
        sl.append(["rank", "pump", "method", "eff_pct", "Q/BEP", "npsh_margin_m",
                   "head_margin_pct", "score", "notes"])
        for c in sl[1]:
            c.font = _KEY
        for i, cand in enumerate(res.selection, 1):
            dd = cand.as_dict()
            sl.append([i, dd["pump"], dd["method"], dd["efficiency_pct"], dd["bep_ratio"],
                       dd["npsh_margin_m"], dd["head_margin_pct"], dd["score"],
                       "; ".join(dd["reasons"])])
        _autosize(sl)

    if res.surge is not None:
        sg = wb.create_sheet("Surge")
        for k, v in res.surge.as_dict().items():
            if isinstance(v, (dict, list)):
                sg.append([k, str(v)])
            else:
                sg.append([k, v])
        _autosize(sg)

    if include_report:
        rp = wb.create_sheet("Report")
        for i, line in enumerate(text_report(res).splitlines(), 1):
            rp.cell(i, 1, line)
        rp.column_dimensions["A"].width = 78

    wb.save(path)
    return str(path)


def run_workbook(in_path: str | Path, out_path: str | Path, *, legacy: bool = False) -> ProjectResults:
    data = read_legacy_workbook(in_path) if legacy else read_project(in_path)
    res = Project.from_dict(data).run()
    write_results(res, out_path)
    return res
